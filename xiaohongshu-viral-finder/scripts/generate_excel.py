#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 报告生成器
生成多工作表的分析报告，包含完整数据、选题分布、标题策略和选题建议
"""

import json
from datetime import datetime
from typing import Dict, List, Any
import sys
import os

# 尝试导入 openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl 未安装，将尝试使用 xlsxwriter")


# 如果 openpyxl 不可用，尝试使用 xlsxwriter
if not HAS_OPENPYXL:
    try:
        import xlsxwriter
        HAS_XLSXWRITER = True
    except ImportError:
        HAS_XLSXWRITER = False
        print("[ERROR] 既没有 openpyxl 也没有 xlsxwriter，无法生成 Excel 文件")


# ==================== 样式配置 ====================

if HAS_OPENPYXL:
    # 标题样式
    HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 边框样式
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


# ==================== Excel 生成器 ====================

def generate_excel_report(feeds: List[Dict[str, Any]], analysis: Dict[str, Any], output_path: str) -> str:
    """
    生成 Excel 报告

    Args:
        feeds: 笔记列表（已增强，包含选题分类、标题策略等）
        analysis: 分析结果
        output_path: 输出文件路径

    Returns:
        生成的文件路径
    """
    if not HAS_OPENPYXL and not HAS_XLSXWRITER:
        raise Exception("需要安装 openpyxl 或 xlsxwriter 库")

    if HAS_OPENPYXL:
        return _generate_with_openpyxl(feeds, analysis, output_path)
    else:
        return _generate_with_xlsxwriter(feeds, analysis, output_path)


def _generate_with_openpyxl(feeds: List[Dict[str, Any]], analysis: Dict[str, Any], output_path: str) -> str:
    """使用 openpyxl 生成 Excel 文件"""

    wb = Workbook()

    # 删除默认工作表
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ==================== 工作表 1: 完整数据 ====================
    ws1 = wb.create_sheet("完整数据", 0)

    # 定义列
    columns = [
        ("笔记标题", 50),
        ("选题分类", 12),
        ("标题策略", 20),
        ("标题长度", 10),
        ("点赞", 10),
        ("收藏", 10),
        ("评论", 10),
        ("收藏/点赞比", 12),
        ("爆款指数", 12),
        ("笔记链接", 60)
    ]

    # 写入表头
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws1.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 写入数据
    for row_idx, feed in enumerate(feeds, 2):
        ws1.cell(row=row_idx, column=1, value=feed.get("title", ""))
        ws1.cell(row=row_idx, column=2, value=feed.get("topic", ""))
        ws1.cell(row=row_idx, column=3, value=feed.get("title_strategy", ""))
        ws1.cell(row=row_idx, column=4, value=feed.get("title_length", 0))
        ws1.cell(row=row_idx, column=5, value=feed.get("likes", 0))
        ws1.cell(row=row_idx, column=6, value=feed.get("collects", 0))
        ws1.cell(row=row_idx, column=7, value=feed.get("comments", 0))
        ws1.cell(row=row_idx, column=8, value=feed.get("collect_to_like_ratio", 0))
        ws1.cell(row=row_idx, column=9, value=feed.get("viral_score", 0))

        # 笔记链接
        note_url = f"https://www.xiaohongshu.com/explore/{feed.get('id', '')}"
        ws1.cell(row=row_idx, column=10, value=note_url)

        # 应用边框
        for col_idx in range(1, 11):
            ws1.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws1.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="center")

    # 冻结首行
    ws1.freeze_panes = "A2"

    # ==================== 工作表 2: 选题分布统计 ====================
    ws2 = wb.create_sheet("选题分布", 1)

    ws2.cell(row=1, column=1, value="选题类型")
    ws2.cell(row=1, column=2, value="笔记数量")
    ws2.cell(row=1, column=3, value="占比")

    # 样式
    for col in range(1, 4):
        cell = ws2.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15

    topic_dist = analysis.get("topic_distribution", {})
    total = sum(topic_dist.values())

    for row_idx, (topic, count) in enumerate(sorted(topic_dist.items(), key=lambda x: x[1], reverse=True), 2):
        ws2.cell(row=row_idx, column=1, value=topic)
        ws2.cell(row=row_idx, column=2, value=count)
        ws2.cell(row=row_idx, column=3, value=f"{count/total*100:.1f}%" if total > 0 else "0%")

    # ==================== 工作表 3: 标题策略分析 ====================
    ws3 = wb.create_sheet("标题策略", 2)

    ws3.cell(row=1, column=1, value="策略类型")
    ws3.cell(row=1, column=2, value="笔记数量")
    ws3.cell(row=1, column=3, value="占比")

    for col in range(1, 4):
        cell = ws3.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 15

    strategy_stats = analysis.get("strategy_stats", {})
    total_strategies = sum(strategy_stats.values())

    for row_idx, (strategy, count) in enumerate(sorted(strategy_stats.items(), key=lambda x: x[1], reverse=True), 2):
        ws3.cell(row=row_idx, column=1, value=strategy)
        ws3.cell(row=row_idx, column=2, value=count)
        ws3.cell(row=row_idx, column=3, value=f"{count/total_strategies*100:.1f}%" if total_strategies > 0 else "0%")

    # ==================== 工作表 4: 选题建议 ====================
    ws4 = wb.create_sheet("选题建议", 3)

    ws4.cell(row=1, column=1, value="建议标题")
    ws4.cell(row=1, column=2, value="选题类型")
    ws4.cell(row=1, column=3, value="目标人群")
    ws4.cell(row=1, column=4, value="核心价值")
    ws4.cell(row=1, column=5, value="内容要点")
    ws4.cell(row=1, column=6, value="参考标题")

    for col in range(1, 7):
        cell = ws4.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    ws4.column_dimensions["A"].width = 50
    ws4.column_dimensions["B"].width = 15
    ws4.column_dimensions["C"].width = 20
    ws4.column_dimensions["D"].width = 30
    ws4.column_dimensions["E"].width = 40
    ws4.column_dimensions["F"].width = 50

    suggestions = analysis.get("suggestions", [])
    for row_idx, suggestion in enumerate(suggestions, 2):
        ws4.cell(row=row_idx, column=1, value=suggestion.get("title", ""))
        ws4.cell(row=row_idx, column=2, value=suggestion.get("topic_type", ""))
        ws4.cell(row=row_idx, column=3, value=suggestion.get("target_audience", ""))
        ws4.cell(row=row_idx, column=4, value=suggestion.get("core_value", ""))
        ws4.cell(row=row_idx, column=5, value="\n".join(suggestion.get("content_points", [])))
        ws4.cell(row=row_idx, column=6, value="\n".join(suggestion.get("recommended_titles", [])))

    # ==================== 保存文件 ====================
    wb.save(output_path)
    return output_path


def _generate_with_xlsxwriter(feeds: List[Dict[str, Any]], analysis: Dict[str, Any], output_path: str) -> str:
    """使用 xlsxwriter 生成 Excel 文件（备用方案）"""

    workbook = xlsxwriter.Workbook(output_path)

    # ==================== 工作表 1: 完整数据 ====================
    ws1 = workbook.add_worksheet("完整数据")

    # 格式
    header_format = workbook.add_format({
        'bold': True,
        'font_size': 12,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })

    cell_format = workbook.add_format({
        'valign': 'vcenter',
        'border': 1
    })

    # 列宽
    ws1.set_column('A:A', 50)  # 笔记标题
    ws1.set_column('B:B', 12)  # 选题分类
    ws1.set_column('C:C', 20)  # 标题策略
    ws1.set_column('D:D', 10)  # 标题长度
    ws1.set_column('E:I', 10)  # 数据列
    ws1.set_column('J:J', 60)  # 笔记链接

    # 表头
    headers = ["笔记标题", "选题分类", "标题策略", "标题长度", "点赞", "收藏", "评论",
               "收藏/点赞比", "爆款指数", "笔记链接"]

    for col_idx, header in enumerate(headers):
        ws1.write(0, col_idx, header, header_format)

    # 数据
    for row_idx, feed in enumerate(feeds, 1):
        ws1.write(row_idx, 0, feed.get("title", ""), cell_format)
        ws1.write(row_idx, 1, feed.get("topic", ""), cell_format)
        ws1.write(row_idx, 2, feed.get("title_strategy", ""), cell_format)
        ws1.write(row_idx, 3, feed.get("title_length", 0), cell_format)
        ws1.write(row_idx, 4, feed.get("likes", 0), cell_format)
        ws1.write(row_idx, 5, feed.get("collects", 0), cell_format)
        ws1.write(row_idx, 6, feed.get("comments", 0), cell_format)
        ws1.write(row_idx, 7, feed.get("collect_to_like_ratio", 0), cell_format)
        ws1.write(row_idx, 8, feed.get("viral_score", 0), cell_format)
        ws1.write(row_idx, 9, f"https://www.xiaohongshu.com/explore/{feed.get('id', '')}", cell_format)

    # 冻结首行
    ws1.freeze_panes(1, 0)

    # ==================== 工作表 2-4: 其他统计表 ====================
    # (省略详细代码，结构与 openpyxl 版本类似)

    workbook.close()
    return output_path


# ==================== 主程序入口 ====================

if __name__ == "__main__":
    # 测试代码
    print("Excel 报告生成器已就绪喵～")
    print("这个模块应该被主程序导入使用")
