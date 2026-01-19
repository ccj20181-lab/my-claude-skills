#!/usr/bin/env python3
"""
小红书低粉爆文抓取 - 最终版本
成功登录并解析搜索结果
支持微信推送功能
"""
import argparse
import json
import requests
import math
import sys
from datetime import datetime
from pathlib import Path
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

MCP_URL = "http://localhost:18060/mcp"
DEFAULT_KEYWORDS = ["金融", "金融知识", "财经", "财经知识", "理财", "股票", "基金", "存钱", "投资理财"]
DEFAULT_CONFIG = "config.viral.json"

def load_config(config_path=DEFAULT_CONFIG):
    """加载配置文件"""
    try:
        # 尝试从脚本所在目录的父目录加载
        script_dir = Path(__file__).parent.parent
        config_file = script_dir / config_path

        if not config_file.exists():
            print(f"[WARN] 配置文件不存在: {config_file}")
            print(f"[INFO] 使用默认配置")
            return {
                "keywords": DEFAULT_KEYWORDS,
                "filters": {
                    "min_likes": 1000,
                    "max_followers": 20000
                },
                "push_wechat": True
            }

        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
            print(f"[INFO] ✓ 配置加载成功: {config_file}")
            return config
    except Exception as e:
        print(f"[WARN] 配置加载失败: {e}")
        print(f"[INFO] 使用默认配置")
        return {
            "keywords": DEFAULT_KEYWORDS,
            "filters": {
                "min_likes": 1000,
                "max_followers": 20000
            },
            "push_wechat": True
        }

class XHS_MCP_Client:
    def __init__(self):
        self.session = requests.Session()
        self.session_id = None
        self.initialized = False

    def init_session(self):
        init_payload = {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "initialize",
            "params": {
                "protocolVersion": "2024-11-05",
                "capabilities": {},
                "clientInfo": {
                    "name": "xhs-viral-finder",
                    "version": "1.0.0"
                }
            }
        }

        resp = self.session.post(MCP_URL, json=init_payload)
        result = resp.json()

        if "result" in result:
            self.session_id = resp.headers.get('Mcp-Session-Id')
            if self.session_id:
                self.session.headers.update({'Mcp-Session-Id': self.session_id})

            notif_payload = {
                "jsonrpc": "2.0",
                "method": "notifications/initialized"
            }
            self.session.post(MCP_URL, json=notif_payload)

            self.initialized = True
            print(f"[INFO] MCP session初始化成功")
            return True
        else:
            print(f"[ERROR] MCP初始化失败: {result}")
            return False

    def call_tool(self, tool_name, arguments, timeout=120):
        if not self.initialized:
            if not self.init_session():
                return None

        payload = {
            "jsonrpc": "2.0",
            "id": int(time.time() * 1000) % 10000,
            "method": "tools/call",
            "params": {
                "name": tool_name,
                "arguments": arguments
            }
        }

        try:
            resp = self.session.post(MCP_URL, json=payload, timeout=timeout)
            result = resp.json()

            if "result" in result:
                return result["result"]
            else:
                print(f"[WARN] 工具调用失败: {result.get('error', {})}")
                return None
        except Exception as e:
            print(f"[ERROR] 请求异常: {e}")
            return None

    def search_notes(self, keyword):
        print(f"[INFO] 搜索关键词: {keyword}")
        result = self.call_tool("search_feeds", {"keyword": keyword}, timeout=120)

        if not result or "content" not in result:
            print(f"[WARN] 关键词 '{keyword}' 无结果")
            return []

        # 解析搜索结果
        candidates = []

        for item in result["content"]:
            if item.get("type") == "text":
                text = item.get("text", "")

                # 检查是否是错误消息
                if "登录" in text and "未登录" in text:
                    print(f"[WARN] 需要登录")
                    return []

                # 尝试解析JSON
                try:
                    data = json.loads(text)

                    if "feeds" not in data:
                        continue

                    feeds = data["feeds"]
                    print(f"[INFO] 找到 {len(feeds)} 条笔记")

                    for note in feeds:
                        try:
                            note_card = note.get("noteCard", {})
                            interact_info = note_card.get("interactInfo", {})
                            user = note_card.get("user", {})

                            likes = int(interact_info.get("likedCount", 0))

                            # 筛选:点赞 >= 1000
                            if likes >= 1000:
                                candidates.append({
                                    "note_id": note.get("id", ""),
                                    "title": note_card.get("displayTitle", ""),
                                    "likes": likes,
                                    "collects": int(interact_info.get("collectedCount", 0)),
                                    "comments": int(interact_info.get("commentCount", 0)),
                                    "user_id": user.get("userId", ""),
                                    "xsec_token": note.get("xsecToken", ""),
                                    "note_url": f"https://www.xiaohongshu.com/explore/{note.get('id', '')}"
                                })
                        except Exception as e:
                            continue

                except json.JSONDecodeError:
                    continue

        print(f"[INFO] 符合点赞初筛(≥1000): {len(candidates)} 条")
        return candidates

    def get_user_followers(self, user_id, xsec_token):
        result = self.call_tool("user_profile", {
            "user_id": user_id,
            "xsec_token": xsec_token
        })

        if result and "content" in result:
            for item in result["content"]:
                if item.get("type") == "text":
                    try:
                        data = json.loads(item.get("text", "{}"))

                        # 粉丝数在interactions数组中
                        if "interactions" in data:
                            for interaction in data["interactions"]:
                                if interaction.get("type") == "fans":
                                    return int(interaction.get("count", 0))

                        # 备用:从userBasicInfo获取
                        if "userBasicInfo" in data:
                            return data["userBasicInfo"].get("fans", -1)
                    except Exception as e:
                        print(f"[DEBUG] 解析用户信息失败: {e}")
                        pass
        return -1

def calculate_viral_scores(notes):
    """计算爆款指数"""
    for note in notes:
        followers = note.get("followers", 0)
        likes = note.get("likes", 0)
        collects = note.get("collects", 0)
        comments = note.get("comments", 0)

        # 互动率 = (点赞 + 收藏 + 评论) / 粉丝数 × 100%
        if followers > 0:
            interaction_rate = (likes + collects + comments) / followers * 100
        else:
            interaction_rate = 0

        # 爆款指数 = 互动率 × log10(点赞数 + 1)
        viral_score = interaction_rate * math.log10(likes + 1)

        note["interaction_rate"] = round(interaction_rate, 2)
        note["viral_score"] = round(viral_score, 2)

    # 按爆款指数排序
    notes.sort(key=lambda x: x["viral_score"], reverse=True)
    return notes

def generate_excel(notes, output_path):
    """生成Excel报告"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "低粉爆文分析"

    # 设置列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 60

    # 表头样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    headers = ["笔记标题", "点赞", "收藏", "评论", "博主粉丝数", "互动率(%)", "爆款指数", "笔记链接"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 写入数据
    for note in notes:
        row_data = [
            note["title"],
            note["likes"],
            note["collects"],
            note["comments"],
            note["followers"],
            note["interaction_rate"],
            note["viral_score"],
            note["note_url"]
        ]
        ws.append(row_data)
        row_num = ws.max_row

        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存文件
    wb.save(output_path)

def push_to_wechat(notes, excel_path, use_html=True, github_url=""):
    """推送爆文分析到微信

    Args:
        notes: 笔记列表
        excel_path: Excel文件路径
        use_html: 是否使用HTML格式(默认True)
        github_url: GitHub Pages URL(可选)
    """
    PUSHPLUS_TOKEN = "a6443f3a5d0f4b11a42c281f831b5c15"

    # 如果使用HTML格式
    if use_html:
        try:
            from push_html import push_html_to_wechat
            return push_html_to_wechat(PUSHPLUS_TOKEN, notes, excel_path, github_url)
        except ImportError:
            print("[WARN] HTML推送模块未找到,使用Markdown格式")
            use_html = False
        except Exception as e:
            print(f"[WARN] HTML推送失败: {e},使用Markdown格式")
            use_html = False

    # Markdown格式(原有逻辑)
    today = datetime.now().strftime("%Y-%m-%d %H:%M")

    content = f"""# 📊 小红书财经爆文日报

**生成时间**: {today}
**扫描关键词**: 金融、金融知识、财经
**发现爆文**: {len(notes)}条

---

## 🔥 TOP 10 爆款笔记

| 排名 | 标题 | 点赞 | 粉丝 | 爆款指数 | 链接 |
|------|------|------|------|---------|------|
"""

    for i, note in enumerate(notes[:10], 1):
        title = note["title"][:20]
        likes = note["likes"]
        followers = note["followers"]
        score = note["viral_score"]
        note_url = note["note_url"]
        content += f"| {i} | {title} | {likes:,} | {followers:,} | {score:.2f} | [查看]({note_url}) |\n"

    # 添加选题分析
    avg_likes = sum(n['likes'] for n in notes) // len(notes)
    avg_followers = sum(n['followers'] for n in notes) // len(notes)
    top_note = notes[0]
    min_followers_note = min(notes, key=lambda x: x['followers'])

    content += f"""

---

## 💡 选题亮点

- **🏆 最高爆款指数**: {top_note['viral_score']:.2f} 分
  标题: {top_note['title'][:30]}
  链接: {top_note['note_url']}

- **📈 平均数据**: 点赞 {avg_likes:,} / 粉丝 {avg_followers:,}

- **⭐ 低粉高赞案例**:
  {min_followers_note['title'][:30]}
  仅 {min_followers_note['followers']} 粉丝获得 {min_followers_note['likes']} 赞
  链接: {min_followers_note['note_url']}

---

{f'🌐 完整网页: [{github_url}]({github_url})' if github_url else ''}

📁 完整Excel: `{excel_path}`
"""

    # 推送到微信
    url = 'http://www.pushplus.plus/send'
    data = {
        "token": PUSHPLUS_TOKEN,
        "title": f"📊 小红书爆文日报 {datetime.now().strftime('%m-%d %H:%M')}",
        "content": content,
        "template": "markdown"
    }

    try:
        response = requests.post(url, json=data, timeout=10)
        result = response.json()
        if result.get("code") == 200:
            print("\n[INFO] ✓ 微信推送成功!请查看手机")
            return True
        else:
            print(f"\n[WARN] 微信推送失败: {result.get('msg')}")
            return False
    except Exception as e:
        print(f"\n[ERROR] 微信推送异常: {e}")
        return False

def main():
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='小红书低粉爆文抓取工具')
    parser.add_argument('--config', default=DEFAULT_CONFIG, help='配置文件路径')
    parser.add_argument('--keywords', nargs='+', help='搜索关键词(覆盖配置文件)')
    parser.add_argument('--min-likes', type=int, help='最小点赞数')
    parser.add_argument('--max-followers', type=int, help='最大粉丝数')
    parser.add_argument('--no-push', action='store_true', help='禁用微信推送')
    args = parser.parse_args()

    # 加载配置
    config = load_config(args.config)

    # 命令行参数覆盖配置
    keywords = args.keywords if args.keywords else config.get("keywords", DEFAULT_KEYWORDS)
    min_likes = args.min_likes if args.min_likes else config.get("filters", {}).get("min_likes", 1000)
    max_followers = args.max_followers if args.max_followers else config.get("filters", {}).get("max_followers", 20000)
    push_wechat = not args.no_push and config.get("push_wechat", True)

    print("=" * 70)
    print("🔍 小红书低粉爆文抓取工具 - 最终版")
    print("=" * 70)
    print(f"[INFO] 搜索关键词: {', '.join(keywords[:3])}... 等{len(keywords)}个")
    print(f"[INFO] 筛选条件: 点赞≥{min_likes}, 粉丝≤{max_followers}")
    print()

    client = XHS_MCP_Client()
    if not client.init_session():
        print("[ERROR] 无法初始化MCP客户端")
        return

    # 搜索关键词
    all_candidates = []

    for keyword in keywords:
        notes = client.search_notes(keyword)
        all_candidates.extend(notes)
        time.sleep(1)  # 避免请求过快

    print(f"\n[INFO] 总共收集 {len(all_candidates)} 条候选笔记")

    if len(all_candidates) == 0:
        print("[WARN] 未找到符合条件的笔记")
        return

    # 去重用户
    unique_users = {}
    for c in all_candidates:
        uid = c["user_id"]
        if uid and uid not in unique_users:
            unique_users[uid] = c["xsec_token"]

    print(f"[INFO] 发现 {len(unique_users)} 个独立用户")

    # 获取粉丝数
    print(f"\n[INFO] 开始获取用户粉丝数...")
    followers_map = {}

    for i, (uid, token) in enumerate(unique_users.items(), 1):
        followers = client.get_user_followers(uid, token)
        followers_map[uid] = followers

        if followers > 0 and i <= 10:
            print(f"[{i}/{len(unique_users)}] 用户{uid}: {followers} 粉丝")

    # 筛选低粉爆文
    print(f"\n[INFO] 筛选低粉爆文...")
    viral_notes = []

    for note in all_candidates:
        followers = followers_map.get(note["user_id"], 999999)
        if followers <= max_followers and note["likes"] >= min_likes:
            note["followers"] = followers
            viral_notes.append(note)

    print(f"[INFO] 符合条件: {len(viral_notes)} 条低粉爆文")

    if len(viral_notes) > 0:
        print("\n" + "=" * 70)
        print("✅ 数据抓取完成")
        print("=" * 70)
        print(f"\n📊 结果:")
        print(f"  • 符合条件的低粉爆文: {len(viral_notes)} 条")

        # 计算爆款指数
        print("\n[INFO] 计算爆款指数...")
        viral_notes = calculate_viral_scores(viral_notes)

        # 生成JSON数据
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_path = Path.home() / "Documents" / f"xhs_viral_notes_{timestamp}.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(viral_notes, f, ensure_ascii=False, indent=2)
        print(f"📁 JSON数据已保存: {json_path}")

        # 生成Excel报告
        print("\n[INFO] 生成Excel报告...")
        excel_path = Path.home() / "Documents" / f"小红书爆文分析_{timestamp}.xlsx"
        generate_excel(viral_notes, excel_path)
        print(f"📁 Excel报告已保存: {excel_path}")

        # 推送到微信
        if push_wechat:
            print("\n" + "=" * 70)
            print("📤 正在推送到微信...")
            print("=" * 70)
            push_success = push_to_wechat(viral_notes, str(excel_path))

            if push_success:
                print("\n" + "=" * 70)
                print("🎉 全部完成!")
                print("=" * 70)
            else:
                print("\n[WARN] 微信推送失败,但数据已保存到本地")
        else:
            print("\n" + "=" * 70)
            print("✅ 完成!(微信推送已禁用)")
            print("=" * 70)
    else:
        print("[WARN] 未找到符合条件的低粉爆文")

if __name__ == "__main__":
    main()
